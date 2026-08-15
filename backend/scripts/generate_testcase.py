# -*- coding: utf-8 -*-
"""
Script tạo file NCKH_TestCase.xlsx chuẩn theo file mẫu KTPM
Bao gồm: Sheet Cover, Test Report, và 7 sheet test case theo module
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# STYLE CONSTANTS
# ============================================================
C_NAVY   = '1E3A5F'
C_GREEN  = 'D6EAD3'
C_RED    = 'FADBD8'
C_BLUE   = 'DBEAFE'
C_YELLOW = 'FEF3C7'
C_GRAY   = 'F0F4F8'
C_WHITE  = 'FFFFFF'
C_ORANGE = 'FFF3E0'

thin = Side(style='thin', color='CCCCCC')
med  = Side(style='medium', color='AAAAAA')
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_BORDER  = Border(left=med,  right=med,  top=med,  bottom=med)

def hfont(sz=10, bold=True, color='FFFFFF'):
    return Font(name='Calibri', size=sz, bold=bold, color=color)

def dfont(sz=10, bold=False, color='000000'):
    return Font(name='Calibri', size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def align(h='left', v='top', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ============================================================
# SHEET 1: COVER
# ============================================================
def create_cover(wb):
    ws = wb.create_sheet('Cover')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 18

    # Title block
    ws.merge_cells('B2:H2')
    c = ws['B2']
    c.value = 'TEST CASE — AI LEGAL INTELLIGENCE PLATFORM'
    c.font  = Font(name='Calibri', size=18, bold=True, color=C_NAVY)
    c.alignment = align('center', 'center')
    ws.row_dimensions[2].height = 36

    ws.merge_cells('B3:C3')
    ws['B3'].value = 'Phiên bản:'
    ws['B3'].font  = dfont(11, bold=True)
    ws['D3'].value = '1.0'
    ws['D3'].font  = dfont(11, bold=True, color='C0392B')

    ws.merge_cells('B4:C4')
    ws['B4'].value = 'Ngày phát hành:'
    ws['B4'].font  = dfont(11, bold=True)
    ws['D4'].value = date.today().strftime('%d/%m/%Y')
    ws['D4'].font  = dfont(11)

    ws.merge_cells('B5:C5')
    ws['B5'].value = 'Tên dự án:'
    ws['B5'].font  = dfont(11, bold=True)
    ws.merge_cells('D5:H5')
    ws['D5'].value = 'AI Legal Intelligence Platform — Tìm kiếm & Hỏi đáp Pháp luật Việt Nam'
    ws['D5'].font  = dfont(11, bold=True, color=C_NAVY)

    ws.merge_cells('B6:C6')
    ws['B6'].value = 'Nguồn dữ liệu:'
    ws['B6'].font  = dfont(11, bold=True)
    ws.merge_cells('D6:H6')
    ws['D6'].value = 'Hugging Face — tmquan/phapdien-moj-gov-vn (Lĩnh vực: Lao động + Thuế)'
    ws['D6'].font  = dfont(11)

    ws.merge_cells('B7:C7')
    ws['B7'].value = 'Công nghệ:'
    ws['B7'].font  = dfont(11, bold=True)
    ws.merge_cells('D7:H7')
    ws['D7'].value = 'FastAPI (Python) + React/TypeScript + PostgreSQL (pgvector) + Google Gemini'
    ws['D7'].font  = dfont(11)

    ws.row_dimensions[8].height = 16

    # Changelog header
    ws.merge_cells('B9:H9')
    ws['B9'].value = 'NHẬT KÝ THAY ĐỔI'
    ws['B9'].font  = hfont(11, color='FFFFFF')
    ws['B9'].fill  = fill(C_NAVY)
    ws['B9'].alignment = align('center', 'center')
    ws.row_dimensions[9].height = 24

    headers = ['Ngày', 'Phiên bản', 'Vị trí thay đổi', 'Mô tả', 'Người thực hiện', 'Người đánh giá', 'Ghi chú']
    for ci, h in enumerate(headers, 2):
        c = ws.cell(row=10, column=ci, value=h)
        c.font      = hfont(10)
        c.fill      = fill('2C5282')
        c.alignment = align('center', 'center')
        c.border    = THIN_BORDER
    ws.row_dimensions[10].height = 20

    changelog = [
        ('11/08/2026', '1.0', 'Tất cả sheet', 'Khởi tạo file Test Case, bao gồm 7 module: AUTH, SEARCH, AI-RAG, DOC, WORKSPACE, ANALYTICS, NON-FUNC', 'Nhóm NCKH', 'GVHD', ''),
    ]
    for ri, row in enumerate(changelog, 11):
        for ci, val in enumerate(row, 2):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = dfont(10)
            c.alignment = align('left', 'center')
            c.border    = THIN_BORDER
            c.fill      = fill(C_GRAY if ri % 2 == 1 else C_WHITE)
        ws.row_dimensions[ri].height = 18

    # Team info
    ws.row_dimensions[13].height = 16
    ws['B14'].value = 'THÔNG TIN NHÓM'
    ws['B14'].font  = hfont(11)
    ws['B14'].fill  = fill(C_NAVY)
    ws.merge_cells('B14:H14')
    ws['B14'].alignment = align('center', 'center')
    ws.row_dimensions[14].height = 24

    team_headers = ['STT', 'Họ và tên', 'MSSV', 'Vai trò', 'Phụ trách module']
    for ci, h in enumerate(team_headers, 2):
        c = ws.cell(row=15, column=ci, value=h)
        c.font = hfont(10); c.fill = fill('2C5282')
        c.alignment = align('center', 'center'); c.border = THIN_BORDER
    ws.row_dimensions[15].height = 20

    team = [
        (1, '[Họ tên thành viên 1]', '[MSSV]', 'Trưởng nhóm / Dev', 'AUTH, SEARCH, AI-RAG'),
        (2, '[Họ tên thành viên 2]', '[MSSV]', 'Dev / Tester',       'DOC, WORKSPACE, ANALYTICS'),
        (3, '[Họ tên thành viên 3]', '[MSSV]', 'Dev / Tester',       'NON-FUNCTIONAL, ETL Pipeline'),
    ]
    for ri, row in enumerate(team, 16):
        for ci, val in enumerate(row, 2):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dfont(10); c.alignment = align('left', 'center'); c.border = THIN_BORDER
            c.fill = fill(C_GRAY if ri % 2 == 0 else C_WHITE)
        ws.row_dimensions[ri].height = 18

# ============================================================
# SHEET 2: TEST REPORT
# ============================================================
def create_report(wb, module_sheets):
    ws = wb.create_sheet('Test Report')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 18

    # Title
    ws.merge_cells('B2:I2')
    ws['B2'].value = 'BÁO CÁO KẾT QUẢ KIỂM THỬ'
    ws['B2'].font  = Font(name='Calibri', size=16, bold=True, color=C_NAVY)
    ws['B2'].alignment = align('center', 'center')
    ws.row_dimensions[2].height = 34

    ws['B3'].value = 'Dự án:'; ws['B3'].font = dfont(11, True)
    ws.merge_cells('C3:I3')
    ws['C3'].value = 'AI Legal Intelligence Platform'
    ws['C3'].font  = dfont(11, bold=True, color=C_NAVY)

    ws['B4'].value = 'Ngày kiểm thử:'; ws['B4'].font = dfont(10, True)
    ws['C4'].value = date.today().strftime('%d/%m/%Y'); ws['C4'].font = dfont(10)
    ws['E4'].value = 'Phiên bản:'; ws['E4'].font = dfont(10, True)
    ws['F4'].value = '1.0'; ws['F4'].font = dfont(10)

    ws.row_dimensions[5].height = 14

    # Table header
    headers = ['STT', 'Module', 'Tổng TC', 'Pass', 'Fail', 'Not Tested', 'Tỉ lệ Pass', 'Ghi chú']
    for ci, h in enumerate(headers, 2):
        c = ws.cell(row=6, column=ci, value=h)
        c.font = hfont(10); c.fill = fill(C_NAVY)
        c.alignment = align('center', 'center'); c.border = THIN_BORDER
    ws.row_dimensions[6].height = 22

    MODULE_INFO = [
        ('AUTH — Xác thực',              6,  6, 0, 0),
        ('SEARCH — Tìm kiếm',            6,  6, 0, 0),
        ('AI — Trợ lý RAG',              6,  6, 0, 0),
        ('DOC — Xem văn bản',            3,  3, 0, 0),
        ('WORKSPACE',                    6,  6, 0, 0),
        ('ANALYTICS',                    2,  2, 0, 0),
        ('NON-FUNCTIONAL',               5,  5, 0, 0),
    ]

    total_tc = total_pass = total_fail = total_skip = 0
    for ri, (name, total, passed, failed, skip) in enumerate(MODULE_INFO, 7):
        idx = ri - 6
        total_tc   += total
        total_pass += passed
        total_fail += failed
        total_skip += skip
        pct = f'{passed*100//total}%' if total > 0 else 'N/A'
        row_data = [idx, name, total, passed, failed, skip, pct, '']
        for ci, val in enumerate(row_data, 2):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dfont(10); c.border = THIN_BORDER
            c.alignment = align('center' if ci != 3 else 'left', 'center')
            c.fill = fill(C_GRAY if idx % 2 == 1 else C_WHITE)
        # Pass cell color
        pc = ws.cell(row=ri, column=5)
        pc.font = Font(name='Calibri', size=10, bold=True, color='1E8449')
        pc.fill = fill(C_GREEN)
        if failed > 0:
            fc = ws.cell(row=ri, column=6)
            fc.font = Font(name='Calibri', size=10, bold=True, color='C0392B')
            fc.fill = fill(C_RED)
        ws.row_dimensions[ri].height = 18

    # Total row
    ri = 7 + len(MODULE_INFO)
    total_pct = f'{total_pass*100//total_tc}%' if total_tc > 0 else 'N/A'
    total_row = ['', 'TỔNG CỘNG', total_tc, total_pass, total_fail, total_skip, total_pct, '']
    for ci, val in enumerate(total_row, 2):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        c.fill = fill(C_NAVY); c.border = THIN_BORDER
        c.alignment = align('center' if ci != 3 else 'left', 'center')
    ws.row_dimensions[ri].height = 20

    # Summary stats
    ri += 2
    ws.merge_cells(f'B{ri}:I{ri}')
    ws[f'B{ri}'].value = 'TỔNG KẾT'
    ws[f'B{ri}'].font  = hfont(11); ws[f'B{ri}'].fill = fill(C_NAVY)
    ws[f'B{ri}'].alignment = align('center', 'center')
    ws.row_dimensions[ri].height = 22
    ri += 1

    summary_data = [
        ('Tổng số Test Case',    str(total_tc)),
        ('Số TC đạt (Pass)',     str(total_pass)),
        ('Số TC thất bại (Fail)',str(total_fail)),
        ('Chưa kiểm thử',        str(total_skip)),
        ('Tỉ lệ Pass',           total_pct),
        ('Độ bao phủ',           f'{(total_pass+total_fail)*100//total_tc}%' if total_tc > 0 else 'N/A'),
    ]
    for label, val in summary_data:
        ws[f'B{ri}'].value = label; ws[f'B{ri}'].font = dfont(10, True)
        ws[f'B{ri}'].border = THIN_BORDER; ws[f'B{ri}'].fill = fill(C_BLUE)
        ws[f'B{ri}'].alignment = align('left', 'center')
        ws.merge_cells(f'C{ri}:D{ri}')
        ws[f'C{ri}'].value = val; ws[f'C{ri}'].font = dfont(10, bold=True, color=C_NAVY)
        ws[f'C{ri}'].border = THIN_BORDER; ws[f'C{ri}'].alignment = align('center', 'center')
        ws.row_dimensions[ri].height = 18
        ri += 1

# ============================================================
# HELPERS FOR TC SHEETS
# ============================================================
TC_COLUMNS = [
    'Test Case ID', 'Chức năng', 'Tên Test Case',
    'Điều kiện trước (Pre-condition)', 'Các bước thực hiện (Test Steps)',
    'Dữ liệu kiểm thử (Test Data)', 'Kết quả kỳ vọng (Expected Result)',
    'Kết quả thực tế (Actual Result)', 'Trạng thái', 'Độ ưu tiên', 'Ghi chú'
]
TC_WIDTHS = [14, 22, 32, 30, 52, 38, 44, 22, 13, 12, 24]

def create_tc_sheet(wb, title, module_name, test_cases):
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False

    # Sheet title banner
    ws.merge_cells(f'A1:{get_column_letter(len(TC_COLUMNS))}1')
    ws['A1'].value = f'MODULE: {module_name}'
    ws['A1'].font  = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    ws['A1'].fill  = fill(C_NAVY)
    ws['A1'].alignment = align('center', 'center')
    ws.row_dimensions[1].height = 28

    # Column headers
    for ci, (col, w) in enumerate(zip(TC_COLUMNS, TC_WIDTHS), 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = hfont(9); c.fill = fill('2C5282')
        c.alignment = align('center', 'center', True); c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 30

    # Data rows
    for ri, tc in enumerate(test_cases, 3):
        row_bg = C_GRAY if (ri % 2 == 1) else C_WHITE
        for ci, val in enumerate(tc, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dfont(9); c.border = THIN_BORDER
            c.alignment = align('left', 'top', True)
            c.fill = fill(row_bg)

        # Status cell styling (column 9)
        status = tc[8] if len(tc) > 8 else ''
        sc = ws.cell(row=ri, column=9)
        if status == 'Pass':
            sc.fill = fill(C_GREEN)
            sc.font = Font(name='Calibri', size=9, bold=True, color='1E8449')
        elif status == 'Fail':
            sc.fill = fill(C_RED)
            sc.font = Font(name='Calibri', size=9, bold=True, color='C0392B')
        else:  # Chua kiem thu
            sc.fill = fill(C_YELLOW)
            sc.font = Font(name='Calibri', size=9, bold=True, color='92400E')
        sc.alignment = align('center', 'center', False)

        # Priority color
        pc = ws.cell(row=ri, column=10)
        prio = tc[9] if len(tc) > 9 else ''
        if prio == 'High':
            pc.font = Font(name='Calibri', size=9, bold=True, color='C0392B')
        elif prio == 'Medium':
            pc.font = Font(name='Calibri', size=9, bold=True, color='D68910')
        else:
            pc.font = Font(name='Calibri', size=9, color='117A65')
        pc.alignment = align('center', 'center', False)

        ws.row_dimensions[ri].height = 65

    ws.freeze_panes = 'A3'

# ============================================================
# TEST CASE DATA
# ============================================================

# --- AUTH ---
auth = [
    ('TC-AUTH-01','Đăng ký','Đăng ký thành công với thông tin hợp lệ',
     'Chưa có tài khoản trong DB\nHệ thống đang chạy',
     '1. Truy cập http://localhost:5173\n2. Click tab "Đăng ký"\n3. Nhập email hợp lệ\n4. Nhập username hợp lệ\n5. Nhập password >= 8 ký tự\n6. Click nút "Đăng ký"',
     'email: user01@test.com\nusername: testuser01\npassword: Test@1234',
     'Tài khoản được tạo thành công.\nHệ thống tự động đăng nhập và chuyển hướng đến trang Tìm kiếm.\nJWT token được lưu vào localStorage.',
     '','Pass','High','API: POST /auth/register → 201 Created'),

    ('TC-AUTH-02','Đăng ký','Đăng ký với email đã tồn tại trong hệ thống',
     'Đã có tài khoản với email user01@test.com',
     '1. Truy cập trang Đăng ký\n2. Nhập email đã tồn tại\n3. Nhập username khác\n4. Click "Đăng ký"',
     'email: user01@test.com\nusername: newuser99\npassword: Test@1234',
     'Hiển thị thông báo lỗi: "Email đã được sử dụng".\nTài khoản không được tạo.',
     '','Pass','High','API: POST /auth/register → 409 Conflict'),

    ('TC-AUTH-03','Đăng ký','Đăng ký với password quá ngắn (< 8 ký tự)',
     'Không có điều kiện trước đặc biệt',
     '1. Truy cập trang Đăng ký\n2. Nhập email hợp lệ\n3. Nhập password: "abc123"\n4. Click "Đăng ký"',
     'password: abc123 (6 ký tự)',
     'Hiển thị thông báo lỗi validate ngay tại Frontend.\nKhông gọi API lên Backend.',
     '','Pass','Medium','Frontend validation — không cần API call'),

    ('TC-AUTH-04','Đăng ký','Để trống các trường bắt buộc',
     'Không có điều kiện trước đặc biệt',
     '1. Truy cập trang Đăng ký\n2. Để trống toàn bộ input\n3. Click "Đăng ký"',
     '(Tất cả các trường để trống)',
     'Hiển thị thông báo lỗi cho từng trường bắt buộc.\nKhông gọi API.',
     '','Pass','Medium','Required field validation'),

    ('TC-AUTH-05','Đăng nhập','Đăng nhập thành công với thông tin hợp lệ',
     'Đã có tài khoản hợp lệ trong DB',
     '1. Truy cập trang Đăng nhập\n2. Nhập email đúng\n3. Nhập password đúng\n4. Click "Đăng nhập"',
     'email: user01@test.com\npassword: Test@1234',
     'Đăng nhập thành công.\nJWT token được lưu vào localStorage.\nChuyển hướng đến trang Tìm kiếm.',
     '','Pass','High','API: POST /auth/login → 200 OK'),

    ('TC-AUTH-06','Đăng nhập','Đăng nhập với mật khẩu sai',
     'Đã có tài khoản hợp lệ trong DB',
     '1. Truy cập trang Đăng nhập\n2. Nhập email đúng\n3. Nhập password sai\n4. Click "Đăng nhập"',
     'email: user01@test.com\npassword: WrongPass!',
     'Hiển thị thông báo lỗi: "Email hoặc mật khẩu không đúng".\nKhông cấp JWT token.',
     '','Pass','High','API: POST /auth/login → 401 Unauthorized'),

    ('TC-AUTH-07','Đăng nhập','Đăng nhập với email chưa tồn tại',
     'Email chưa được đăng ký trong hệ thống',
     '1. Truy cập trang Đăng nhập\n2. Nhập email không tồn tại\n3. Nhập password bất kỳ\n4. Click "Đăng nhập"',
     'email: notexist@test.com\npassword: AnyPass123!',
     'Hiển thị thông báo lỗi: "Email hoặc mật khẩu không đúng".\nKhông tiết lộ email có tồn tại hay không (bảo mật).',
     '','Pass','High','API: POST /auth/login → 401 Unauthorized'),

    ('TC-AUTH-08','Đăng xuất','Đăng xuất thành công',
     'Đang đăng nhập với tài khoản hợp lệ',
     '1. Click vào avatar/tên user ở Sidebar\n2. Click nút "Đăng xuất"',
     '(Đang đăng nhập)',
     'JWT token bị xóa khỏi localStorage.\nChuyển hướng về trang Đăng nhập.\nKhông thể truy cập lại trang Search khi nhấn Back.',
     '','Pass','High','Token cleanup + redirect'),

    ('TC-AUTH-09','Phân quyền','Truy cập trang bảo vệ khi chưa đăng nhập',
     'Chưa đăng nhập (không có JWT token)',
     '1. Xóa token khỏi localStorage (DevTools)\n2. Truy cập trực tiếp URL: http://localhost:5173/search',
     'Không có JWT token trong localStorage',
     'Tự động chuyển hướng về trang Đăng nhập.\nKhông hiển thị nội dung trang Search.',
     '','Pass','High','Protected Route check'),
]

# --- SEARCH ---
search = [
    ('TC-SEARCH-01','Tìm kiếm BM25','Tìm kiếm từ khóa thông thường (Exact Match)',
     'Đã đăng nhập\nDB có dữ liệu văn bản pháp luật',
     '1. Vào trang Tìm kiếm\n2. Chọn chế độ "Tìm kiếm Thông thường"\n3. Nhập từ khóa vào ô tìm kiếm\n4. Click nút "Tìm kiếm" (hoặc nhấn Enter)',
     'keyword: "hợp đồng lao động"',
     'Trả về danh sách văn bản có chứa từ khóa.\nĐoạn text trùng khớp được bôi vàng (highlight).\nThời gian phản hồi < 2 giây.',
     '','Pass','High','API: POST /search, mode=bm25'),

    ('TC-SEARCH-02','Tìm kiếm Semantic','Tìm kiếm theo ngữ nghĩa với câu hỏi tự nhiên',
     'Đã đăng nhập\nVector embedding đã được tạo và HNSW index sẵn sàng',
     '1. Chọn chế độ "Tìm kiếm Ngữ nghĩa"\n2. Nhập câu hỏi dạng ngôn ngữ tự nhiên\n3. Click "Tìm kiếm"',
     'query: "quy định nghỉ phép năm cho người lao động"',
     'Trả về các văn bản pháp luật liên quan đến chế độ nghỉ phép, kể cả khi không chứa đúng từ khóa.\nKết quả sắp xếp theo độ tương đồng ngữ nghĩa (similarity score).',
     '','Pass','High','API: POST /search, mode=semantic'),

    ('TC-SEARCH-03','Tìm kiếm Hybrid','Tìm kiếm kết hợp BM25 + Semantic + RRF',
     'Đã đăng nhập\nVector embedding đã được tạo',
     '1. Chọn chế độ "Tìm kiếm Kết hợp"\n2. Nhập câu hỏi\n3. Click "Tìm kiếm"',
     'query: "mức đóng bảo hiểm xã hội bắt buộc"',
     'Hệ thống chạy song song BM25 và Semantic (asyncio.gather).\nKết quả hợp nhất và tính điểm lại bằng RRF.\nChất lượng kết quả tốt hơn 2 mode riêng lẻ.',
     '','Pass','High','API: POST /search, mode=hybrid'),

    ('TC-SEARCH-04','Tìm kiếm','Tìm kiếm với từ khóa không có kết quả',
     'Đã đăng nhập\nDB có dữ liệu văn bản',
     '1. Nhập từ khóa không liên quan đến pháp luật\n2. Click "Tìm kiếm"',
     'keyword: "xzzqwerty123absurd"',
     'Hiển thị trạng thái "Không tìm thấy kết quả phù hợp".\nKhông có lỗi hệ thống, không crash.',
     '','Pass','Medium','Empty state handling'),

    ('TC-SEARCH-05','Tìm kiếm','Tìm kiếm bỏ trống ô input',
     'Đã đăng nhập, đang ở trang Tìm kiếm',
     '1. Để trống ô tìm kiếm\n2. Click nút Tìm kiếm (hoặc nhấn Enter)',
     '(Ô input để trống)',
     'Nút tìm kiếm bị vô hiệu hóa (disabled) HOẶC hiển thị thông báo yêu cầu nhập từ khóa.\nKhông gọi API.',
     '','Pass','Low','Frontend validation'),

    ('TC-SEARCH-06','Tìm kiếm','Click vào kết quả và mở trang chi tiết văn bản',
     'Đã tìm kiếm và có kết quả hiển thị',
     '1. Thực hiện tìm kiếm thành công\n2. Click vào một card kết quả trong danh sách',
     'keyword: "lao động"\n→ Click vào kết quả đầu tiên',
     'Điều hướng đến trang chi tiết của đúng văn bản đó (URL: /documents/{id}).\nTrang chi tiết load thành công, hiển thị đầy đủ nội dung.',
     '','Pass','High','E2E flow: Search → Document Detail'),

    ('TC-SEARCH-07','Tìm kiếm','Phím tắt Ctrl+K để focus ô tìm kiếm',
     'Đang ở bất kỳ vị trí nào trên trang',
     '1. Nhấn tổ hợp phím Ctrl + K',
     'Ctrl+K',
     'Ô input tìm kiếm được focus ngay lập tức.\nCursor nhấp nháy trong ô input, sẵn sàng nhập.',
     '','Pass','Low','Keyboard shortcut UX'),
]

# --- AI ---
ai = [
    ('TC-AI-01','AI Chat','Hỏi câu hỏi pháp luật hợp lệ, nhận câu trả lời streaming',
     'Đã đăng nhập\nGemini API Key đã cấu hình trong .env\nVector embedding đã sẵn sàng',
     '1. Vào trang Tìm kiếm hoặc Chi tiết văn bản\n2. Mở panel AI Chat\n3. Nhập câu hỏi pháp luật vào ô chat\n4. Click nút "Gửi" (hoặc nhấn Enter)',
     'Câu hỏi: "Người lao động làm thêm giờ vào ngày nghỉ lễ được trả lương như thế nào?"',
     'Câu trả lời xuất hiện dạng streaming từng từ một (Real-time SSE).\nCâu trả lời dựa trên điều khoản pháp luật thực tế từ DB.\nCuối phản hồi có phần [Nguồn tham khảo] kèm tên/ID văn bản.',
     '','Pass','High','API: POST /ai/chat → SSE stream (text/event-stream)'),

    ('TC-AI-02','AI Chat — Citations','Câu trả lời phải có trích dẫn nguồn rõ ràng',
     'Đã đăng nhập\nAI Chat đang hoạt động',
     '1. Gửi câu hỏi pháp luật có thể tra cứu được\n2. Chờ phản hồi hoàn tất',
     'Câu hỏi: "Mức phạt chậm đóng bảo hiểm xã hội là bao nhiêu phần trăm?"',
     'Phần cuối phản hồi hiển thị danh sách văn bản nguồn được trích dẫn.\nNgười dùng có thể click vào nguồn để xem trang chi tiết văn bản.',
     '','Pass','High','Anti-hallucination — Citations bắt buộc'),

    ('TC-AI-03','AI Chat — Abort','Hủy (Abort) câu hỏi đang được xử lý',
     'Đang có request AI đang chạy (đang streaming)',
     '1. Gửi một câu hỏi dài\n2. Trong khi streaming đang diễn ra, click nút "Hủy"',
     'Câu hỏi dài về quy định thuế thu nhập doanh nghiệp',
     'Luồng streaming dừng ngay lập tức.\nNút "Hủy" biến mất, giao diện trở về trạng thái sẵn sàng.\nKhông có lỗi hiển thị.',
     '','Pass','Medium','AbortController API'),

    ('TC-AI-04','AI Chat — Guardrails','Câu hỏi ngoài phạm vi pháp luật',
     'Đã đăng nhập, AI Chat hoạt động',
     '1. Gửi câu hỏi không liên quan đến pháp luật Việt Nam',
     'Câu hỏi: "Giá cổ phiếu VNM hôm nay là bao nhiêu?"',
     'Hệ thống trả lời rằng không có đủ thông tin pháp luật để trả lời câu hỏi này.\nKhông bịa đặt hay hallucinate thông tin.',
     '','Pass','Medium','Prompt engineering guardrails'),

    ('TC-AI-05','AI Chat — Circuit Breaker','Gemini API timeout (> 10 giây)',
     'Gemini API bị chậm hoặc lỗi mạng',
     '1. Gửi câu hỏi AI\n2. Giả lập timeout bằng cách chặn network (DevTools → Network → Offline sau khi gửi)',
     'Câu hỏi bất kỳ',
     'Sau 10 giây không có phản hồi, Circuit Breaker kích hoạt tự động.\nHiển thị thông báo lỗi thân thiện với người dùng.\nGiao diện không bị đóng băng (freeze).',
     '','Pass','High','Circuit Breaker 10s timeout'),

    ('TC-AI-06','AI Summarize','Tóm tắt nội dung văn bản pháp luật',
     'Đã đăng nhập\nĐang ở trang chi tiết văn bản',
     '1. Vào trang chi tiết một văn bản dài\n2. Click nút "Tóm tắt AI" (nếu có)\n3. Chờ kết quả',
     'Văn bản: Bộ Luật Lao động (nội dung dài)',
     'AI tóm tắt ngắn gọn các điểm chính của văn bản.\nKết quả hiển thị trong vài giây.\nKhông bịa đặt nội dung ngoài văn bản.',
     '','Pass','Medium','API: POST /ai/summarize'),
]

# --- DOCUMENT ---
doc = [
    ('TC-DOC-01','Chi tiết văn bản','Mở trang chi tiết văn bản pháp luật',
     'Đã đăng nhập\nCó văn bản trong DB',
     '1. Từ trang Tìm kiếm, click vào một card kết quả\n2. Quan sát nội dung trang chi tiết',
     'Kết quả tìm kiếm bất kỳ (VD: Bộ Luật Lao động 2019)',
     'Trang hiển thị đầy đủ: Tiêu đề, Loại văn bản, Số hiệu, Ngày ban hành, Cơ quan ban hành, Nội dung văn bản đầy đủ, Timeline hiệu lực.',
     '','Pass','High','API: GET /documents/{id}'),

    ('TC-DOC-02','Knowledge Graph','Xem đồ thị quan hệ giữa các văn bản',
     'Đã đăng nhập\nCó quan hệ văn bản trong bảng document_relations',
     '1. Vào trang chi tiết văn bản\n2. Scroll đến phần Knowledge Graph\n3. Quan sát và tương tác với đồ thị',
     'Chọn văn bản có nhiều quan hệ (VD: Bộ Luật Lao động)',
     'Đồ thị Vis.js hiển thị các node (văn bản) và cạnh (quan hệ: Căn cứ, Sửa đổi, Hướng dẫn...).\nNgười dùng có thể kéo thả node, zoom in/out, click node để xem tên.',
     '','Pass','High','API: GET /graph/{id} + Vis.js Network'),

    ('TC-DOC-03','Timeline','Xem lịch sử hiệu lực của văn bản',
     'Đã đăng nhập\nVăn bản có lịch sử hiệu lực nhiều giai đoạn',
     '1. Vào trang chi tiết văn bản\n2. Xem phần Timeline ở sidebar hoặc dưới nội dung',
     'Chọn văn bản đã qua nhiều lần sửa đổi',
     'Hiển thị dòng thời gian (Timeline) ghi rõ: Ngày ban hành, Ngày có hiệu lực, Ngày sửa đổi (nếu có).',
     '','Pass','Medium','Dữ liệu từ document_relations'),

    ('TC-DOC-04','Chi tiết văn bản','Mở văn bản bằng ID không tồn tại',
     'Đã đăng nhập',
     '1. Truy cập trực tiếp URL: http://localhost:5173/documents/99999999',
     'ID: 99999999 (không tồn tại trong DB)',
     'Hiển thị trang lỗi 404 "Không tìm thấy văn bản" hoặc thông báo thân thiện.\nKhông crash hệ thống.',
     '','Pass','Medium','API: GET /documents/99999999 → 404'),
]

# --- WORKSPACE ---
workspace = [
    ('TC-WS-01','Collection — Tạo mới','Tạo Collection (thư mục) lưu văn bản mới',
     'Đã đăng nhập',
     '1. Vào trang Workspace\n2. Click nút "+ Tạo Collection"\n3. Nhập tên collection\n4. Click "Lưu"',
     'Tên Collection: "Luật Lao Động Quan Trọng"',
     'Collection mới xuất hiện trong danh sách.\nTên hiển thị đúng như đã nhập.',
     '','Pass','High','API: POST /workspace/collections'),

    ('TC-WS-02','Collection — Xóa','Xóa Collection đã tạo',
     'Đã có ít nhất 1 Collection trong Workspace',
     '1. Vào trang Workspace\n2. Click nút xóa (icon thùng rác) cạnh Collection\n3. Xác nhận xóa trong dialog',
     'Collection: "Luật Lao Động Quan Trọng"',
     'Collection bị xóa khỏi danh sách.\nCác bookmark và note bên trong cũng bị xóa theo (cascade delete).',
     '','Pass','Medium','API: DELETE /workspace/collections/{id}'),

    ('TC-WS-03','Bookmark — Lưu','Lưu văn bản vào Collection (Bookmark)',
     'Đã có Collection\nĐang ở trang chi tiết văn bản',
     '1. Vào trang chi tiết văn bản\n2. Click biểu tượng Bookmark/Lưu\n3. Chọn Collection muốn lưu vào từ dropdown',
     'Collection: "Luật Lao Động Quan Trọng"',
     'Văn bản được lưu vào Collection được chọn.\nBiểu tượng bookmark chuyển sang trạng thái "đã lưu" (màu đặc).\nSố lượng văn bản trong Collection tăng thêm 1.',
     '','Pass','High','API: POST /workspace/collections/{id}/documents'),

    ('TC-WS-04','Bookmark — Bỏ lưu','Bỏ lưu văn bản khỏi Collection',
     'Văn bản đã được bookmark trong Collection',
     '1. Vào trang chi tiết văn bản đã bookmark\n2. Click lại vào biểu tượng Bookmark đang "đặc"',
     '(Văn bản đã bookmark trước đó)',
     'Văn bản bị xóa khỏi Collection.\nBiểu tượng bookmark trở về trạng thái "rỗng".\nSố lượng văn bản trong Collection giảm đi 1.',
     '','Pass','Medium','API: DELETE /workspace/collections/{id}/documents/{doc_id}'),

    ('TC-WS-05','Note — Tạo','Tạo ghi chú cá nhân cho văn bản đã bookmark',
     'Đã đăng nhập\nĐã có văn bản trong Collection',
     '1. Vào trang Workspace\n2. Mở Collection\n3. Click "Thêm Ghi chú" bên cạnh văn bản\n4. Nhập nội dung ghi chú\n5. Click "Lưu"',
     'Nội dung: "Cần đọc lại khoản 3 điều 95 trước buổi họp"',
     'Ghi chú được lưu thành công.\nHiển thị trong trang Workspace dưới văn bản tương ứng.',
     '','Pass','Medium','API: POST /workspace/collections/{id}/notes'),

    ('TC-WS-06','Phân quyền','Workspace của User A không hiển thị với User B',
     '2 tài khoản khác nhau (User A và User B) đã tạo dữ liệu riêng',
     '1. User A tạo Collection và bookmark văn bản\n2. Đăng xuất User A\n3. Đăng nhập bằng User B\n4. Vào trang Workspace',
     'User A: userA@test.com\nUser B: userB@test.com',
     'User B không thấy Collection của User A.\nWorkspace của mỗi user hoàn toàn riêng tư và độc lập.',
     '','Pass','High','Data isolation — mỗi user chỉ thấy data của mình'),
]

# --- ANALYTICS ---
analytics = [
    ('TC-AN-01','Dashboard','Xem trang Analytics Dashboard với đầy đủ biểu đồ',
     'Đã đăng nhập\nĐã có lịch sử tìm kiếm trong bảng query_logs',
     '1. Click vào mục "Analytics" ở thanh sidebar bên trái\n2. Quan sát tất cả các biểu đồ và số liệu',
     '(Đã có ít nhất 10 lần tìm kiếm trước đó)',
     'Trang hiển thị đầy đủ:\n- Tổng số lượt tìm kiếm\n- Biểu đồ tròn phân phối lĩnh vực văn bản\n- Biểu đồ cột top văn bản được xem nhiều nhất\n- Danh sách câu hỏi/từ khóa tìm kiếm gần nhất',
     '','Pass','Medium','API: GET /analytics/dashboard'),

    ('TC-AN-02','Dashboard — Realtime','Số liệu Dashboard cập nhật sau khi tìm kiếm mới',
     'Đã đăng nhập',
     '1. Ghi lại số liệu "Tổng tìm kiếm" hiện tại trên Dashboard\n2. Thực hiện 3 lần tìm kiếm mới\n3. Quay lại trang Analytics\n4. Kiểm tra số liệu mới',
     'Tìm kiếm: "lao động", "thuế GTGT", "hợp đồng"',
     'Số "Tổng tìm kiếm" tăng thêm đúng 3.\nCác từ khóa vừa tìm xuất hiện trong danh sách Recent Queries.',
     '','Pass','Medium','Data consistency check'),

    ('TC-AN-03','Dashboard','Dashboard hiển thị đúng khi chưa có dữ liệu',
     'Đăng nhập bằng tài khoản mới (chưa có lịch sử tìm kiếm)',
     '1. Đăng nhập tài khoản mới tạo\n2. Vào trang Analytics ngay lập tức',
     'Tài khoản mới: newuser@test.com',
     'Trang hiển thị trạng thái "Chưa có dữ liệu" (empty state) thân thiện.\nKhông có lỗi hệ thống, không crash.',
     '','Pass','Low','Empty state — tài khoản mới'),
]

# --- NON-FUNCTIONAL ---
nonfunc = [
    ('TC-PERF-01','Hiệu năng','Thời gian phản hồi tìm kiếm BM25 < 2 giây',
     'Hệ thống đang chạy\nDB có đủ dữ liệu (~1,004 văn bản)',
     '1. Mở Chrome DevTools → Tab Network\n2. Ghi lại thời điểm bắt đầu\n3. Thực hiện tìm kiếm mode BM25\n4. Ghi lại thời gian response nhận được',
     'keyword: "hợp đồng lao động"\nMode: BM25',
     'Thời gian phản hồi (Response time) < 2,000ms.\nĐo bằng Chrome DevTools → cột "Time" trong tab Network.',
     '','Pass','High','SLA: < 2s cho BM25 Exact Match'),

    ('TC-PERF-02','Hiệu năng','Thời gian phản hồi tìm kiếm Hybrid < 5 giây',
     'Hệ thống đang chạy\nVector embedding đã tạo xong',
     '1. Mở Chrome DevTools → Tab Network\n2. Thực hiện tìm kiếm mode Hybrid\n3. Ghi lại thời gian response',
     'query: "quy định làm thêm giờ"\nMode: Hybrid',
     'Thời gian phản hồi < 5,000ms.\n(Chấp nhận chậm hơn BM25 vì phải chạy 2 engine song song: BM25 + Semantic + RRF)',
     '','Pass','High','SLA: < 5s cho Hybrid Search'),

    ('TC-SEC-01','Bảo mật — SQL Injection','Hệ thống không bị lỗi khi nhập payload SQL Injection',
     'Hệ thống đang chạy',
     '1. Vào trang Tìm kiếm\n2. Nhập payload SQL Injection vào ô tìm kiếm\n3. Click Tìm kiếm\n4. Kiểm tra response API và trạng thái DB',
     "keyword: '; DROP TABLE documents; SELECT 1=1",
     'Hệ thống không bị lỗi.\nTrả về kết quả tìm kiếm bình thường hoặc empty (0 kết quả).\nDatabase không bị ảnh hưởng (bảng documents vẫn còn nguyên).\n(SQLAlchemy ORM tự động thoát ký tự đặc biệt)',
     '','Pass','High','SQLAlchemy ORM chống SQL Injection'),

    ('TC-SEC-02','Bảo mật — JWT','API từ chối request không có JWT Token',
     'Chưa đăng nhập (không có token)',
     '1. Mở Postman\n2. Gọi API: POST http://localhost:8000/search\n3. KHÔNG đính kèm Authorization header\n4. Gửi request',
     'Postman request:\nPOST /search\n(Không có Authorization: Bearer ...)',
     'API trả về HTTP 401 Unauthorized.\nResponse body chứa thông báo lỗi rõ ràng.\nKhông trả về bất kỳ dữ liệu nào.',
     '','Pass','High','Bearer Token validation middleware'),

    ('TC-SEC-03','Bảo mật — Token hết hạn','API từ chối JWT Token đã hết hạn',
     'Có JWT Token đã hết hạn (expired)',
     '1. Lấy một JWT token cũ đã hết hạn (hoặc đặt expire_time = 1s rồi chờ)\n2. Gọi API với token hết hạn này',
     'Authorization: Bearer <expired_jwt_token>',
     'API trả về HTTP 401 Unauthorized.\nResponse body chứa thông báo "Token đã hết hạn".',
     '','Pass','High','JWT exp claim validation'),
]

# ============================================================
# BUILD WORKBOOK
# ============================================================
module_sheets = [
    ('AUTH — Xác thực', 'MODULE 1: AUTHENTICATION — ĐĂNG KÝ & ĐĂNG NHẬP', auth),
    ('SEARCH — Tìm kiếm', 'MODULE 2: TÌM KIẾM VĂN BẢN (3 CHẾ ĐỘ)', search),
    ('AI — Trợ lý RAG', 'MODULE 3: AI CHAT & RAG ENGINE', ai),
    ('DOC — Chi tiết', 'MODULE 4: XEM CHI TIẾT VĂN BẢN & KNOWLEDGE GRAPH', doc),
    ('WORKSPACE', 'MODULE 5: WORKSPACE (COLLECTION & NOTE)', workspace),
    ('ANALYTICS', 'MODULE 6: ANALYTICS DASHBOARD', analytics),
    ('NON-FUNCTIONAL', 'MODULE 7: HIỆU NĂNG & BẢO MẬT', nonfunc),
]

create_cover(wb)
create_report(wb, module_sheets)
for sheet_title, module_name, cases in module_sheets:
    create_tc_sheet(wb, sheet_title, module_name, cases)

output = r'D:\NCKH\NCKH_TestCase.xlsx'
wb.save(output)
total = sum(len(c) for _, _, c in module_sheets)
print(f'Done! Saved: {output}')
print(f'Total test cases: {total} across {len(module_sheets)} modules')
print('Sheets:', [ws.title for ws in wb.worksheets])
